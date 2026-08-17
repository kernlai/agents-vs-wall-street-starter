from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


def norm(value) -> str:
    return "" if value is None else str(value).strip().casefold()


def main() -> None:
    parser = argparse.ArgumentParser(description="Write three forecasts into one supplied workbook template")
    parser.add_argument("--company", required=True)
    parser.add_argument("--forecast", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--companies", default="challenge/companies.json")
    args = parser.parse_args()
    registry = json.loads(Path(args.companies).read_text(encoding="utf-8"))["companies"]
    company = next(item for item in registry if item["ticker"].split(":")[-1] == args.company)
    payload = json.loads(Path(args.forecast).read_text(encoding="utf-8"))
    values = {item["metric"]: item["value"] for item in payload["forecasts"]}
    expected = [item["label"] for item in company["metrics"]]
    if set(values) != set(expected):
        raise ValueError("forecast metrics do not exactly match challenge configuration")
    destination = Path(args.output)
    workbook = openpyxl.load_workbook(args.template)
    sheet = workbook["Summary"]
    header = next((row for row in range(1, 31) if norm(sheet.cell(row, 1).value) == "metric" and norm(sheet.cell(row, 2).value) == "units" and norm(sheet.cell(row, 3).value) == norm(company["period"])), None)
    if header is None:
        raise ValueError("workbook header was not found")
    for offset, metric in enumerate(company["metrics"], 1):
        row = header + offset
        if norm(sheet.cell(row, 1).value) != norm(metric["label"]) or norm(sheet.cell(row, 2).value) != norm(metric["units"]):
            raise ValueError(f"template contract mismatch at row {row}")
        value = values[metric["label"]]
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError(f"non-numeric forecast for {metric['label']}")
        sheet.cell(row, 3).value = float(value)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    print(json.dumps({"company": args.company, "forecasts": 3, "output": str(destination)}))


if __name__ == "__main__":
    main()
