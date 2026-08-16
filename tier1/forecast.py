#!/usr/bin/env python3
"""Tier 1 forecaster — seasonal naive with drift.

Run from the repo root:

    python3 tier1/forecast.py

Reads:
    challenge/companies.json      (targets, labels, units, output filenames)
    tier1/history.json            (verified historical figures you filled in)
    challenge/templates/*.xlsx    (untouched OpenStocks templates)

Writes:
    submission/<TICKER>-<PERIOD>.xlsx   (three numbers per workbook, nothing else)
    logs/tier1-run-<timestamp>.log      (full reasoning trail)

Methods
-------
growth      forecast = same-period-last-year x (1 + g), where g blends the most
            recent reported YoY growth (weight 0.7) with last year's same-period
            YoY growth (weight 0.3). Used for currency amounts and EPS.
pct_blend   forecast = 0.6 x latest reported value + 0.4 x same-period-last-year.
            Used for margins / comp-sales style metrics, which mean-revert and
            should never be extrapolated multiplicatively.

The script refuses to write a workbook if the minimum history for any of its
metrics is missing — better a loud failure at 15:00 than a silent zero at 18:00.
"""

from __future__ import annotations

import json
import re
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COMPANIES = ROOT / "challenge" / "companies.json"
HISTORY = Path(__file__).resolve().parent / "history.json"
TEMPLATES = ROOT / "challenge" / "templates"
SUBMISSION = ROOT / "submission"
LOGS = ROOT / "logs"

RECENT_W = 0.7   # weight on most recent reported YoY growth
LATEST_W = 0.6   # pct_blend weight on the most recent reported value
GROWTH_CAP = 0.30  # |g| above this is almost certainly a data error


# --------------------------------------------------------------------------
# Fiscal-period handling. FY2026Q3 -> (2026, "Q", 3); FY2026H1 -> (2026, "H", 1);
# FY2026 -> (2026, "Y", 0). Ordering key lets "latest reported" mean something.
# --------------------------------------------------------------------------
PERIOD_RE = re.compile(r"^FY(\d{4})(?:(Q|H)(\d))?$")


def parse_period(label: str) -> tuple[int, str, int]:
    m = PERIOD_RE.match(label.strip())
    if not m:
        raise ValueError(f"Unrecognised period label: {label!r}")
    year = int(m.group(1))
    kind = m.group(2) or "Y"
    idx = int(m.group(3) or 0)
    return year, kind, idx


def order_key(label: str) -> tuple[int, float]:
    year, kind, idx = parse_period(label)
    frac = {"Y": 1.0, "H": idx / 2, "Q": idx / 4}[kind]
    return year, frac


def year_earlier(label: str) -> str:
    year, kind, idx = parse_period(label)
    suffix = "" if kind == "Y" else f"{kind}{idx}"
    return f"FY{year - 1}{suffix}"


# --------------------------------------------------------------------------
# Forecasting methods
# --------------------------------------------------------------------------
class MissingHistory(Exception):
    pass


def clean_series(raw: dict) -> dict[str, float]:
    out = {}
    for period, entry in raw.items():
        value = entry.get("value") if isinstance(entry, dict) else entry
        if value is not None:
            out[period] = float(value)
    return out


def forecast_growth(target: str, series: dict[str, float], log: list[str]) -> float:
    prior = year_earlier(target)          # same period last year
    prior2 = year_earlier(prior)          # same period two years back
    if prior not in series:
        raise MissingHistory(f"need {prior} (same period last year)")

    growths: list[tuple[float, float, str]] = []  # (weight, g, explanation)

    # For annual targets, an interim pair (e.g. FY2026H1 vs FY2025H1) is the
    # freshest evidence available — it already covers part of the target year —
    # so it takes priority as the "recent" component.
    if parse_period(target)[1] == "Y":
        halves = [p for p in series if parse_period(p)[1] == "H"]
        for p in sorted(halves, key=order_key, reverse=True):
            twin = year_earlier(p)
            if twin in series and series[twin] != 0:
                g = series[p] / series[twin] - 1
                growths.append((RECENT_W, g, f"interim YoY {twin}->{p}: {g:+.1%}"))
                break

    # Otherwise: most recent reported period, of the same kind as the target,
    # that has its prior-year twin. Skip the seasonal pair itself (p == prior)
    # so the same evidence is never counted twice.
    if not growths:
        same_kind = [p for p in series if parse_period(p)[1] == parse_period(target)[1]]
        for p in sorted(same_kind, key=order_key, reverse=True):
            if p == prior:
                continue
            twin = year_earlier(p)
            if twin in series and series[twin] != 0:
                g = series[p] / series[twin] - 1
                growths.append((RECENT_W, g, f"recent YoY {twin}->{p}: {g:+.1%}"))
                break

    if prior2 in series and series[prior2] != 0:
        g = series[prior] / series[prior2] - 1
        growths.append((0.3, g, f"seasonal YoY {prior2}->{prior}: {g:+.1%}"))

    if not growths:
        raise MissingHistory(
            f"have {prior} but no growth evidence (need {prior2}, or a recent "
            f"period plus its prior-year twin)"
        )

    total_w = sum(w for w, _, _ in growths)
    g = sum(w * gr for w, gr, _ in growths) / total_w
    for _, _, note in growths:
        log.append(f"      {note}")
    if abs(g) > GROWTH_CAP:
        log.append(f"      WARNING blended growth {g:+.1%} is extreme — check the inputs")
    forecast = series[prior] * (1 + g)
    log.append(
        f"      growth method: {prior}={series[prior]:,.2f} x (1{g:+.3f}) "
        f"= {forecast:,.2f}"
    )
    return forecast


def forecast_pct_blend(target: str, series: dict[str, float], log: list[str]) -> float:
    prior = year_earlier(target)
    if prior not in series:
        raise MissingHistory(f"need {prior} (same period last year)")
    latest = max(series, key=order_key)
    if latest == prior:
        log.append(f"      pct_blend: only {prior} available, carrying it forward")
        return series[prior]
    forecast = LATEST_W * series[latest] + (1 - LATEST_W) * series[prior]
    log.append(
        f"      pct_blend: {LATEST_W}x{latest}({series[latest]:.2f}) + "
        f"{1 - LATEST_W:.1f}x{prior}({series[prior]:.2f}) = {forecast:.2f}"
    )
    return forecast


METHODS = {"growth": forecast_growth, "pct_blend": forecast_pct_blend}


def round_for_units(value: float, units: str) -> float:
    if "share" in units or units == "GBp":
        return round(value, 2)   # EPS in dollars / pence
    return round(value, 1)       # amounts in millions, percentage points


# --------------------------------------------------------------------------
# Workbook writing. Mirrors check-forecasts.mjs: find the Metric/Units/period
# header row, verify each label+units, write plain numbers in column 3.
# --------------------------------------------------------------------------
def write_workbook(company: dict, values: dict[str, float], log: list[str]) -> Path:
    template = TEMPLATES / company["outputFile"]
    if not template.exists():
        raise FileNotFoundError(f"template missing: {template}")
    wb = load_workbook(template)
    ws = wb["Summary"]

    header_row = None
    for r in range(1, 31):
        if (
            str(ws.cell(r, 1).value or "").strip() == "Metric"
            and str(ws.cell(r, 2).value or "").strip() == "Units"
            and str(ws.cell(r, 3).value or "").strip() == company["period"]
        ):
            header_row = r
            break
    if header_row is None:
        raise RuntimeError(f"{template.name}: Metric/Units/{company['period']} header not found")

    for i, metric in enumerate(company["metrics"]):
        r = header_row + 1 + i
        label = str(ws.cell(r, 1).value or "").strip()
        units = str(ws.cell(r, 2).value or "").strip()
        if label != metric["label"] or units != metric["units"]:
            raise RuntimeError(
                f"{template.name} row {r}: expected {metric['label']!r}/{metric['units']!r}, "
                f"found {label!r}/{units!r}"
            )
        ws.cell(r, 3).value = values[metric["label"]]
        log.append(f"    Summary!C{r} <- {values[metric['label']]}  ({label}, {units})")

    SUBMISSION.mkdir(exist_ok=True)
    out = SUBMISSION / company["outputFile"]
    wb.save(out)
    return out


# --------------------------------------------------------------------------
def main() -> int:
    companies = json.loads(COMPANIES.read_text())["companies"]
    history = json.loads(HISTORY.read_text())

    log: list[str] = [f"Tier 1 run — {time.strftime('%Y-%m-%d %H:%M:%S %Z')}"]
    failures: list[str] = []

    for company in companies:
        ticker = company["ticker"].split(":")[-1]  # LSE:HAS -> HAS
        log.append(f"\n{company['company']} ({ticker}) -> {company['period']}")
        hist = history.get(ticker)
        if hist is None:
            failures.append(f"{ticker}: no entry in history.json")
            continue
        target = hist["target"]
        if target != company["period"]:
            failures.append(f"{ticker}: history target {target} != {company['period']}")
            continue

        values: dict[str, float] = {}
        company_ok = True
        for metric in company["metrics"]:
            label = metric["label"]
            log.append(f"  {label} [{metric['units']}]")
            spec = hist["metrics"].get(label)
            if spec is None:
                failures.append(f"{ticker} / {label}: missing from history.json")
                company_ok = False
                continue
            series = clean_series(spec["series"])
            override = spec.get("override") or {}
            try:
                if override.get("value") is not None:
                    raw = float(override["value"])
                    log.append(
                        f"      OVERRIDE {raw} — {override.get('source', 'no source given')}"
                    )
                else:
                    raw = METHODS[spec["method"]](target, series, log)
            except MissingHistory as e:
                failures.append(f"{ticker} / {label}: {e}")
                company_ok = False
                continue
            values[label] = round_for_units(raw, metric["units"])
            log.append(f"      => {values[label]}")

        if company_ok:
            out = write_workbook(company, values, log)
            log.append(f"  wrote {out.relative_to(ROOT)}")
        else:
            log.append("  SKIPPED — history incomplete (see failures)")

    if failures:
        log.append("\nFAILURES:")
        log.extend(f"  - {f}" for f in failures)

    LOGS.mkdir(exist_ok=True)
    log_path = LOGS / f"tier1-run-{time.strftime('%Y%m%d-%H%M%S')}.log"
    text = "\n".join(log)
    log_path.write_text(text + "\n")
    print(text)
    print(f"\nlog saved to {log_path.relative_to(ROOT)}")

    if failures:
        print("\nFill the missing figures in tier1/history.json and re-run.")
        return 1
    print("\nAll four workbooks written. Now run: npm run check:submission")
    return 0


if __name__ == "__main__":
    sys.exit(main())