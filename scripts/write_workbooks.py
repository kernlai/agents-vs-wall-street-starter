#!/usr/bin/env python3
"""Write the four submission workbooks from evaluation/forecasts.json.

Copies each supplied template, locates the Summary sheet's header row by matching
"Metric" / "Units" / <fiscal period> exactly as the organisers' checker does, then
writes the three forecasts into the period column. Everything else — sheet names,
labels, units, styling, the yellow fills — is left untouched, because openpyxl
round-trips the workbook rather than rebuilding it.

Refuses to write when anything looks wrong:
  - a metric label or unit in the template that does not match companies.json
  - a forecast that is missing or non-numeric
  - a header row it cannot find
A refusal is deliberate: a workbook silently written to the wrong cell passes the
organisers' check and fails the judging.

    python3 scripts/write_workbooks.py                 # write all four
    python3 scripts/write_workbooks.py --dry-run       # report, write nothing
"""
from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "challenge" / "templates"
SUBMISSION = ROOT / "submission"
COMPANIES = ROOT / "challenge" / "companies.json"
FORECASTS = ROOT / "evaluation" / "forecasts.json"

MAX_HEADER_SCAN = 30  # the checker scans the first 30 rows; match that exactly


def norm(v) -> str:
    return "" if v is None else str(v).strip().casefold()


def find_header_row(ws, period: str) -> int | None:
    """Same rule the organisers' checker uses: row where A=Metric, B=Units, C=<period>."""
    for r in range(1, MAX_HEADER_SCAN + 1):
        if (norm(ws.cell(r, 1).value) == "metric"
                and norm(ws.cell(r, 2).value) == "units"
                and norm(ws.cell(r, 3).value) == norm(period)):
            return r
    return None


def write_one(company: dict, values: dict, dry_run: bool) -> tuple[bool, list[str]]:
    msgs: list[str] = []
    name = company["outputFile"]
    src, dst = TEMPLATES / name, SUBMISSION / name

    if not src.exists():
        return False, [f"template missing: {src}"]

    wb = openpyxl.load_workbook(src)
    if "Summary" not in wb.sheetnames:
        return False, ["template has no 'Summary' sheet"]
    ws = wb["Summary"]

    header = find_header_row(ws, company["period"])
    if header is None:
        return False, [f"no header row matching Metric/Units/{company['period']} in the first {MAX_HEADER_SCAN} rows"]
    msgs.append(f"header row {header}")

    ok = True
    for i, metric in enumerate(company["metrics"]):
        row = header + 1 + i
        label, units = metric["label"], metric["units"]

        got_label = ws.cell(row, 1).value
        got_units = ws.cell(row, 2).value
        if norm(got_label) != norm(label):
            msgs.append(f"row {row}: template label {got_label!r} != companies.json {label!r}")
            ok = False
            continue
        if norm(got_units) != norm(units):
            msgs.append(f"row {row}: template units {got_units!r} != companies.json {units!r}")
            ok = False
            continue

        v = values.get(label)
        if v is None:
            msgs.append(f"row {row}: NO FORECAST for {label!r} — an empty cell scores 5.0")
            ok = False
            continue
        if not isinstance(v, (int, float)) or isinstance(v, bool):
            msgs.append(f"row {row}: forecast for {label!r} is not numeric: {v!r}")
            ok = False
            continue

        if not dry_run:
            ws.cell(row, 3).value = float(v)
        msgs.append(f"C{row} = {float(v):,.4g}   ({label})")

    if ok and not dry_run:
        SUBMISSION.mkdir(exist_ok=True)
        shutil.copyfile(src, dst)          # keep the template byte-identical as the base
        wb.save(dst)
        msgs.append(f"wrote {dst.relative_to(ROOT)}")
    elif not ok:
        msgs.append("REFUSED to write — fix the above first")
    return ok, msgs


def main(argv: list[str]) -> int:
    dry = "--dry-run" in argv
    companies = json.loads(COMPANIES.read_text())["companies"]
    forecasts = json.loads(FORECASTS.read_text())

    print(f"\n  writing submission workbooks{'  (DRY RUN)' if dry else ''}")
    print("  " + "-" * 78)
    all_ok = True
    for c in companies:
        ticker = c["ticker"].split(":")[-1]
        vals = {k: v for k, v in forecasts.get(ticker, {}).items() if not k.startswith("_")}
        ok, msgs = write_one(c, vals, dry)
        all_ok &= ok
        print(f"\n  {c['company']}  ({ticker} · {c['period']})   {'OK' if ok else 'BLOCKED'}")
        for m in msgs:
            print(f"      {m}")
    print("\n  " + "-" * 78)
    print(f"  {'all four workbooks ready' if all_ok else 'NOT ready — see blocked entries above'}\n")
    return 0 if all_ok else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
